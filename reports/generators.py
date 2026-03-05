from datetime import datetime
from io import BytesIO
from typing import Any, List, Optional, Sequence, Tuple
import traceback
from datetime import datetime
from mensajes_logs import logger_
from flask import current_app
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable
from openpyxl.worksheet.table import TableStyleInfo


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.setFont("Helvetica", 8)
            ancho = self._pagesize[0]
            self.drawCentredString(ancho / 2, 0.48 * inch, f"Página {self._pageNumber}/{total_pages}")
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)


def _nombre_empresa() -> str:
    return current_app.config.get("COMPANY_NAME", "Alitas El Comelon")


def _ruta_logo() -> Optional[str]:
    return current_app.config.get("COMPANY_LOGO_PATH")


def _ahora_str() -> str:
    tz = current_app.config.get("REPORTS_TIMEZONE")
    now = datetime.now(tz) if tz else datetime.now()
    return now.strftime("%Y-%m-%d %H:%M:%S")


def _texto(val: Any) -> str:
    if val is None:
        return ""
    try:
        if isinstance(val, (list, tuple, set)):
            parts = [_texto(v) for v in list(val)]
            parts = [p for p in parts if p]
            return ", ".join(parts)
        return str(val)
    except Exception:
        return ""


def _escape_para(s: str) -> str:
    s = str(s)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _separar_camel(s: str) -> str:
    if not s:
        return ""
    out = []
    prev = ""
    for ch in s:
        if prev and ch.isupper() and (prev.islower() or (prev.isupper() and len(out) > 0 and out[-1][-1].islower())):
            out.append(" ")
        out.append(ch)
        prev = ch
    return "".join(out)


def _normalizar_etiqueta_base(s: Any) -> str:
    txt = _texto(s).strip()
    if not txt:
        return ""
    txt = txt.replace("_", " ").replace("-", " ")
    txt = " ".join(txt.split())
    txt = _separar_camel(txt)
    txt = " ".join(txt.split())
    return txt


def _quitar_prefijo_id(label: str) -> str:
    t = (label or "").strip()
    u = t.upper()
    if u.startswith("ID "):
        return t[3:].strip()
    if u.startswith("ID_"):
        return t[3:].strip()
    if u.startswith("ID"):
        rest = t[2:].strip()
        if rest:
            return rest
    return t


def _acento_label_upper(s: str) -> str:
    t = (s or "").strip()
    if not t:
        return ""

    palabras = t.split(" ")
    mapa = {
        "DIRECCION": "DIRECCIÓN",
        "TELEFONO": "TELÉFONO",
        "NUMERO": "NÚMERO",
        "PARAMETRO": "PARÁMETRO",
        "DESCRIPCION": "DESCRIPCIÓN",
        "CANCELACION": "CANCELACIÓN",
        "CREACION": "CREACIÓN",
        "CONFIGURACION": "CONFIGURACIÓN",
        "SUCURSAL": "SUCURSAL",
        "USUARIO": "USUARIO",
        "CLIENTE": "CLIENTE",
        "ENTREGA": "ENTREGA",
        "ORDEN": "ORDEN",
        "FACTURA": "FACTURA",
        "REPARTIDOR": "REPARTIDOR",
        "NOMBRE": "NOMBRE",
        "APELLIDO": "APELLIDO",
        "CODIGO": "CÓDIGO",
        "CATEGORIA": "CATEGORÍA",
        "CANTIDAD": "CANTIDAD",
        "PRECIO": "PRECIO",
        "IMPUESTO": "IMPUESTO",
        "DETALLE": "DETALLE",
        "FECHA": "FECHA",
        "HORA": "HORA",
        "DIA": "DÍA",
        "MES": "MES",
        "ANIO": "AÑO",
        "ANO": "AÑO",
    }

    out = []
    for p in palabras:
        u = p.upper()
        out.append(mapa.get(u, u))
    return " ".join(out)


def _label_columna(s: Any) -> str:
    base = _normalizar_etiqueta_base(s)
    base = _quitar_prefijo_id(base)
    n = " ".join(base.lower().replace("_", " ").split())

    if n in ("us co", "usco", "us co.", "us_co", "direccion entrega", "direccion_entrega"):
        base = "Dirección Entrega"

    if n in ("usuario cliente f", "usuario_cliente_f", "usuario-cliente-f"):
        base = "Usuario Cliente"

    base = base.upper()
    return _acento_label_upper(base)

def _necesita_landscape(cols: int) -> bool:
    try:
        max_cols_portrait = int(current_app.config.get("REPORTS_MAX_COLS_PORTRAIT", 6))
    except Exception:
        max_cols_portrait = 6
    return cols > max_cols_portrait


def _split_threshold() -> int:
    try:
        return int(current_app.config.get("REPORTS_SPLIT_THRESHOLD", 10))
    except Exception:
        return 10


def _max_cols_per_table() -> int:
    try:
        return int(current_app.config.get("REPORTS_MAX_COLS_PER_TABLE", 9))
    except Exception:
        return 9


def _anchor_cols_count() -> int:
    try:
        return int(current_app.config.get("REPORTS_SPLIT_ANCHOR_COLS", 1))
    except Exception:
        return 1


def _should_split(columns_count: int) -> bool:
    return columns_count > _split_threshold()


def _particionar_columnas_con_anclas(
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> List[Tuple[List[str], List[List[Any]]]]:
    cols = list(columns)
    if not cols:
        return [([], [])]

    max_cols = _max_cols_per_table()
    anclas = max(0, min(_anchor_cols_count(), len(cols)))
    if anclas >= len(cols):
        return [(cols, [list(r) for r in rows])]

    ancla_cols = cols[:anclas]
    resto_cols = cols[anclas:]

    chunk_size = max(1, max_cols - anclas)

    out = []
    i = 0
    while i < len(resto_cols):
        chunk_cols = ancla_cols + resto_cols[i : i + chunk_size]
        chunk_rows = []
        for r in rows:
            rr = list(r)
            ancla_vals = rr[:anclas]
            resto_vals = rr[anclas:]
            chunk_rows.append(ancla_vals + resto_vals[i : i + chunk_size])
        out.append((chunk_cols, chunk_rows))
        i += chunk_size

    return out


def _font_sizes_for_cols(ncols: int) -> Tuple[float, float]:
    if ncols >= 10:
        return 7.4, 7.0
    if ncols >= 8:
        return 8.0, 7.4
    return 9.0, 8.2


def _calc_col_widths(avail_width: float, columns: Sequence[str], rows: Sequence[Sequence[Any]]) -> List[float]:
    n = len(columns)
    if n <= 0:
        return []

    sample = rows[:250] if rows else []
    maxlens = []
    for i, c in enumerate(columns):
        m = len(_texto(c))
        for r in sample:
            if i < len(r):
                m = max(m, len(_texto(r[i])))
        maxlens.append(m)

    weights = [max(8, min(70, x)) for x in maxlens]
    total = sum(weights) if weights else 1

    min_w = 1.15 * inch if n >= 8 else 1.25 * inch
    max_w = 3.0 * inch

    widths = []
    for w in weights:
        ww = avail_width * (w / total)
        ww = max(min_w, min(max_w, ww))
        widths.append(ww)

    s = sum(widths)
    if s > 0:
        scale = avail_width / s
        widths = [x * scale for x in widths]

    return widths


def _header_block(report_title: str, printed_by: str) -> Table:
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleCustom",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        spaceAfter=6,
        textColor=colors.black,
    )
    meta_style = ParagraphStyle(
        "MetaCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.grey,
        leading=11,
    )

    text_block = [
        Paragraph(_escape_para(report_title), title_style),
        Paragraph(_escape_para(f"Empresa: {_nombre_empresa()}"), meta_style),
        Paragraph(_escape_para(f"Impreso por: {printed_by}"), meta_style),
        Paragraph(_escape_para(f"Fecha/Hora: {_ahora_str()}"), meta_style),
    ]

    logo_path = _ruta_logo()
    if logo_path:
        try:
            img = RLImage(logo_path)
            target_w = 1.6 * inch
            iw, ih = img.imageWidth, img.imageHeight
            if iw and ih:
                scale = target_w / float(iw)
                img.drawWidth = target_w
                img.drawHeight = float(ih) * scale
            t = Table([[img, text_block]], colWidths=[1.8 * inch, None])
        except Exception:
            t = Table([[text_block]], colWidths=[None])
    else:
        t = Table([[text_block]], colWidths=[None])

    t.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return t


def generar_pdf(report_title: str, columns: Sequence[str], rows: Sequence[Sequence[Any]], printed_by: str) -> bytes:
    try:
        buf = BytesIO()

        cols_fmt = [_label_columna(c) for c in columns]
        page_size = landscape(letter) if _necesita_landscape(len(cols_fmt)) else letter

        left_margin = 0.55 * inch
        right_margin = 0.55 * inch
        top_margin = 0.7 * inch
        bottom_margin = 0.8 * inch

        doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
        title=report_title,
        author=_nombre_empresa(),
        )

        avail_width = page_size[0] - left_margin - right_margin

        story = []
        story.append(_header_block(report_title, printed_by))
        story.append(Spacer(1, 0.18 * inch))

        if _should_split(len(cols_fmt)):
            secciones = _particionar_columnas_con_anclas(cols_fmt, rows)
        else:
            secciones = [(list(cols_fmt), [list(r) for r in rows])]

        for idx, (cols_chunk, rows_chunk) in enumerate(secciones):
            ncols = len(cols_chunk)
            head_fs, body_fs = _font_sizes_for_cols(ncols)

            wrap_header = ParagraphStyle(
                name=f"WrapHeader_{idx}",
                fontName="Helvetica-Bold",
                fontSize=head_fs,
                leading=head_fs + 2,
                alignment=1,
                wordWrap="LTR",
                splitLongWords=0,
                textColor=colors.white,
            )

            wrap_cell = ParagraphStyle(
                name=f"WrapCell_{idx}",
                fontName="Helvetica",
                fontSize=body_fs,
                leading=body_fs + 3,
                wordWrap="LTR",
                splitLongWords=1,
                textColor=colors.black,
            )

            story = []
            story.append(_header_block(report_title, printed_by))
            story.append(Spacer(1, 0.18 * inch))

            if _should_split(len(cols_fmt)):
                secciones = _particionar_columnas_con_anclas(cols_fmt, rows)
            else:
                secciones = [(list(cols_fmt), [list(r) for r in rows])]

            for idx, (cols_chunk, rows_chunk) in enumerate(secciones):
                ncols = len(cols_chunk)
                head_fs, body_fs = _font_sizes_for_cols(ncols)

                wrap_header = ParagraphStyle(
                    name="WrapHeader",
                    fontName="Helvetica-Bold",
                    fontSize=head_fs,
                    leading=head_fs + 2,
                    alignment=1,
                    wordWrap="LTR",
                    splitLongWords=0,
                    textColor=colors.white,
                )

                wrap_cell = ParagraphStyle(
                    name="WrapCell",
                    fontName="Helvetica",
                    fontSize=body_fs,
                    leading=body_fs + 3,
                    wordWrap="LTR",
                    splitLongWords=1,
                    textColor=colors.black,
                )

                header_row = [Paragraph(_escape_para(_texto(c)), wrap_header) for c in cols_chunk]
                data = [header_row]

            for r in rows_chunk:
                fila = []
                for v in r:
                    s = _escape_para(_texto(v)).replace("\n", "<br/>")
                    fila.append(Paragraph(s, wrap_cell))
                data.append(fila)

            col_widths = _calc_col_widths(avail_width, cols_chunk, rows_chunk)
            table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="LEFT")

            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),

                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("VALIGN", (0, 1), (-1, -1), "TOP"),

                ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#9CA3AF")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#F3F4F6")]),

                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]

            table.setStyle(TableStyle(style_cmds))
            story.append(table)

            if idx < len(secciones) - 1:
                story.append(Spacer(1, 0.18 * inch))
                story.append(PageBreak())

        doc.build(story, canvasmaker=NumberedCanvas)
        return buf.getvalue()
    except Exception as error:
        fecha = datetime.now().strftime("%Y%m%d-%H%M%S")
        logger_.Logger.add_to_log("error", str(error), "generar_pdf", fecha)
        logger_.Logger.add_to_log("error", traceback.format_exc(), "generar_pdf", fecha)

def _auto_anchos_excel(ws, columns: Sequence[str], rows: Sequence[Sequence[Any]]):
    widths = [len(_texto(c)) for c in columns]

    for r in rows[:800]:
        for i, v in enumerate(r):
            s = _texto(v)
            if i < len(widths):
                widths[i] = max(widths[i], len(s))

    for i, w in enumerate(widths, start=1):
        val = int(min(max(w + 4, 14), 60))
        ws.column_dimensions[get_column_letter(i)].width = val


def generar_excel(report_title: str, columns: Sequence[str], rows: Sequence[Sequence[Any]], printed_by: str) -> bytes:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        ws.sheet_view.showGridLines = False

        fondo = PatternFill("solid", fgColor="FFFFFF")
        for row in range(1, 160):
            for col in range(1, 80):
                ws.cell(row=row, column=col).fill = fondo

        cols_fmt = [_label_columna(c) for c in columns]

        max_col = max(1, len(cols_fmt))
        last_col = get_column_letter(max(2, max_col))

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 42

        logo_path = _ruta_logo()
        if logo_path:
            try:
                img = XLImage(logo_path)
                size = 90
                img.width = size
                img.height = size
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = size * 0.75
            except Exception:
                ws.row_dimensions[1].height = 65
        else:
            ws.row_dimensions[1].height = 65

        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 10

        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = report_title
        ws["A2"].font = Font(bold=True, name="Calibri", size=16)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

        ws["A3"] = "Empresa:"
        ws["B3"] = _nombre_empresa()
        ws["A4"] = "Impreso por:"
        ws["B4"] = printed_by
        ws["A5"] = "Fecha/Hora:"
        ws["B5"] = _ahora_str()

        for r in range(3, 6):
            ws[f"A{r}"].font = Font(bold=True, name="Calibri", size=11)
            ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")

        start_row = 8

        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        body_font = Font(name="Calibri", size=10)

        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.row_dimensions[start_row].height = 34

        for col_index, col_name in enumerate(cols_fmt, 1):
            cell = ws.cell(row=start_row, column=col_index, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_index, row_data in enumerate(rows, start_row + 1):
            ws.row_dimensions[row_index].height = 18
            for col_index, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_index, column=col_index, value=_texto(value))
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

            for row_index, row_data in enumerate(rows, start_row + 1):
                ws.row_dimensions[row_index].height = 18
                for col_index, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_index, column=col_index, value=_texto(value))
                    cell.font = body_font
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = border

            _auto_anchos_excel(ws, cols_fmt, rows)

            max_row = start_row + len(rows)
            real_last_col = get_column_letter(max_col)

            try:
                ref = f"A{start_row}:{real_last_col}{max_row}"
                tab = XLTable(displayName="TablaReporte", ref=ref)
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
            except Exception:
                pass

            ws.freeze_panes = ws[f"A{start_row + 1}"]

            out = BytesIO()
            wb.save(out)
            return out.getvalue()
    except Exception as error:
        fecha = datetime.now().strftime("%Y%m%d-%H%M%S")
        logger_.Logger.add_to_log("error", str(error), "generar_pdf", fecha)
        logger_.Logger.add_to_log("error", traceback.format_exc(), "generar_pdf", fecha)


def render_pdf(report_title: str, columns: Sequence[str], rows: Sequence[Sequence[Any]], printed_by: str) -> bytes:
    try:
        return generar_pdf(report_title=report_title, columns=columns, rows=rows, printed_by=printed_by)
    except Exception as error:
        fecha = datetime.now().strftime("%Y%m%d-%H%M%S")
        logger_.Logger.add_to_log("error", str(error), "generar_pdf", fecha)
        logger_.Logger.add_to_log("error", traceback.format_exc(), "generar_pdf", fecha)

def render_excel(report_title: str, columns: Sequence[str], rows: Sequence[Sequence[Any]], printed_by: str) -> bytes:
    try:
        return generar_excel(report_title=report_title, columns=columns, rows=rows, printed_by=printed_by)
    except Exception as error:
        fecha = datetime.now().strftime("%Y%m%d-%H%M%S")
        logger_.Logger.add_to_log("error", str(error), "generar_excel", fecha)
        logger_.Logger.add_to_log("error", traceback.format_exc(), "generar_excel", fecha)